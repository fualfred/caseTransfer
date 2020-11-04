#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from xml.dom import minidom


class Utils(object):
    @classmethod
    def get_work_book(cls, path, sheet_index=0):
        try:
            wb = load_workbook(path)
            sheet = wb.worksheets[sheet_index]
            return sheet
        except Exception as e:
            print("please check Excel path or sheet index", e)
            pass

    @classmethod
    def get_max_rows(cls, sheet):
        try:
            return sheet.max_row
        except Exception as e:
            print(e)
            pass

    @classmethod
    def get_max_col(cls, sheet):
        try:
            return sheet.max_column
        except Exception as e:
            print(e)
            pass

    @classmethod
    def get_cell_value(cls, sheet, row, col):
        try:
            cell_value = sheet.cell(row, col).value
            return cell_value
        except Exception as e:
            print(e)
            pass

    @classmethod
    def get_row_value(cls, sheet, row, max_col):
        try:
            values = []
            for i in range(max_col):
                cell_value = cls.get_cell_value(sheet, row, i + 1)
                values.append(str(cell_value))
            return values
        except Exception as e:
            print(e)
            pass

    @classmethod
    def create_dom(cls):
        return minidom.Document()

    @classmethod
    def create_element(cls, doc, tag_name, text=None, attr_value=None, ):
        try:
            elem = doc.createElement(tag_name)
            if attr_value is not None:
                elem.setAttribute("name", attr_value)
            if text is not None:
                text_node = doc.createTextNode(text)
                elem.appendChild(text_node)
            return elem
        except Exception as e:
            print(e)
            pass

    @classmethod
    def get_element_by_attr_name(cls, el, tag_name, attr_value):
        try:
            els = el.getElementsByTagName(tag_name)
            for elem in els:
                attr_real = elem.getAttribute("name")
                if attr_real == attr_value:
                    return elem
            return None
        except Exception as e:
            print(e)
            pass

    @classmethod
    def append_child_elem(cls, parent, child):
        try:
            parent.appendChild(child)
        except Exception as e:
            print(e)
            pass

    @classmethod
    def write_xml(cls, doc, file_name=None):
        try:
            # print(type(xml))
            if file_name is None:
                file_name = "testCase.xml"
            f = open(file_name, 'w', encoding="utf-8")
            doc.writexml(f, encoding="utf-8")
            f.close()
        except Exception as e:
            print(e)
            pass

    @classmethod
    def create_section(cls, doc, tag_name, data):
        try:
            elem = cls.create_element(doc, tag_name, None, None)
            data_section = doc.createCDATASection(data)
            elem.appendChild(data_section)
            return elem
        except Exception as e:
            print(e)

    @classmethod
    def trim_space(cls, str_value):
        if str_value is not None:
            return str_value.strip()
        return None

    @classmethod
    def fill_None_value(cls,msg):
        for i in range(len(msg)):
            if msg[i] == 'None':
                msg[i] = '/'
        return msg