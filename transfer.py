# coding=utf-8
from openpyxl import Workbook, load_workbook
from on_main import run_t

def run(excel_path):
    try:
        wb = Workbook()
        ws=wb.active
        # print(sheetNames)
        test_case_wb = load_workbook(excel_path)
        sheetNames = test_case_wb.sheetnames
        print(sheetNames)
        format_testCases = list()
        for i in range(0,len(sheetNames)):
            test_case_ws=test_case_wb[sheetNames[i]]
            # print(test_case_ws.sheet_state)
            if test_case_ws.sheet_state == 'hidden':
                continue
            m_row = test_case_ws.max_row
            # print(m_row)
        # m_col = ws.max_column
            test_cases = list()
            start_row=9
            for i in range(start_row, m_row+1):
                test_case = list()
                for j in range(1, 8):
                    val = str(test_case_ws.cell(i, j).value)
                    val = val.strip("")
                    test_case.append(val)
                test_cases.append(test_case)
            # print(test_cases)
            first_module = str(test_case_ws.cell(start_row, 1).value)
            second_module = str(test_case_ws.cell(start_row, 2).value)
            for test_case in test_cases:
                format_testCase = list()
                if test_case[0] != "None" and test_case[0] != first_module:
                    first_module = test_case[0]
                    second_module = ""
                if test_case[1] != "None" and test_case[1] != second_module:
                    second_module = test_case[1]
                title = test_case[2]
                if test_case[3] == "None":
                    test_case[3] = 1
                level = int(test_case[3]) + 1
                precondition = test_case[4]
                if precondition == "None":
                    precondition = "/"
                steps = test_case[5]
                expect =test_case[6]

                format_testCase.append(first_module)
                format_testCase.append(second_module)
                format_testCase.append(title)
                format_testCase.append("/")
                format_testCase.append(str(level))
                format_testCase.append(precondition)
                format_testCase.append(steps)
                format_testCase.append(expect)
                
                format_testCases.append(format_testCase)
        # print(format_testCases)   
        sheet_title=["功能模块", "子功能模块", "用例标题", "摘要","用例等级", "前置条件", "测试步骤", "预期结果"]
        # print(format_testCases)
        ws.append(sheet_title)
        for fr_testCase in format_testCases:
            # print(fr_testCase)
            ws.append(fr_testCase)
        # wb.save("testlink.xlsx")
        # print("transfer excel 100%")
        return run_t(wb)
    except Exception as e:
        print(e)
        return "异常"+ str(e)

if __name__ == "__main__":
    run()
